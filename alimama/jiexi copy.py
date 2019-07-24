# !/usr/bin/python3
import json
# import xlwt
# import xlrd
import time
from openpyxl import Workbook

title = [
    "佣金", "优惠券信息", "产品地址", "原价", "卖家昵称", "产品标题", "30天销量", "商品折后价", "月支出佣金",
    "淘客30天月推广量"
]
SHEET_NAME = 'Sheet1'


class ExcelUtil:
    def __init__(self, dirPath, fileName):
        self.dirPath = dirPath
        self.fileName = fileName
        self.fileCount = 1
        detester = time.strftime("%Y-%m-%d", time.localtime())
        self.path = "%s/%s%s_导出数据.xls" % (self.dirPath, self.fileName,
                                          detester)
        self.excelFile = self.createExcel(self.path)  # 创建一个excel

    # def crateExcel2(self):
    #     self.saveExcel()
    #     #创建一个文件
    #     detester = time.strftime("%Y-%m-%d", time.localtime())
    #     self.path = "%s/%s%s_导出数据_%d.xls" % (self.dirPath, self.fileName,
    #                                          detester, self.fileCount)
    #     self.excelFile = self.createExcel(self.path)  # 创建一个excel
    #     self.fileCount += 1
    def readExcelNrows(self):
        try:
            self.saveExcel()
            # book = xlrd.open_workbook(self.path)  # 打开一个excel
            # sheet = book.sheet_by_name(SHEET_NAME)  # 根据顺序获取sheet
            # return sheet.nrows
            #获取所有sheet页名字
            sheet = self.excelFile.get_sheet_by_name(SHEET_NAME)
            row = sheet.nrows
            return row
        except BaseException:
            return 1
            pass
        # return 1

    def readExcel(self, fileName):
        with open(fileName, 'r', encoding='utf8') as fr:
            data = json.load(fr)  # 用json中的load方法，将json串转换成字典
        return data

    def saveExcel(self):
        #detester = time.strftime("%Y-%m-%d", time.localtime())
        # self.excelFile.save(
        #     ("%s%s%s_导出数据.xlsx" % (self.dirPath, self.fileName, detester)))
        # print(self.dirPath)
        try:
            # print(self.path)
            self.excelFile.save(self.path)
        except BaseException:
            pass

    def createExcel(self, fileName):
        book = Workbook()  #创建文件对象
        # book = xlwt.Workbook()  # 创建一个excel对象 只写
        # self.sheet = book.add_sheet(SHEET_NAME,
        #                             cell_overwrite_ok=True)  # 添加一个sheet页
        self.sheet = book.active
        # self.sheet = book.create_sheet(SHEET_NAME)
        for i in range(len(title)):  # 循环列
            self.sheet.cell(1, i+1, title[i])  # 将title数组中的字段写入到0行i列中
        # detester = time.strftime("%Y-%m-%d", time.localtime())

        book.save(fileName)
        return book
        # self.excelFile.save(
        #     ("%s/%s%s_导出数据.xlsx" % (self.dirPath, self.fileName, detester)))

    def writeM(self, items):
        # a = readExcel('testJson.json')
        # print(a)
        # 标题	标题链接	price-info-num	price-info-num1	price-info-text	price-info-num2	价格	box-shop-info	box-shop-right
        # commission_rate => 5000   佣金（5000表示50%，即除以100）
        # coupon_info => 满21.00元减15元      优惠券信息，如无此字段则表示没有优惠券
        # item_url => https://item.ta...?id=543574320825   产品地址
        # reserve_price => 58       原价（未打折前原始标价，不是券前价哦）
        # shop_title => 九州树叶旗舰店          卖家昵称
        # title => 荷叶茶冬瓜荷...天然决明正品子             产品标题
        # volume => 159777       30天销量
        # zk_final_price => 21.9    商品折后价（券前价，就是产品在淘宝上折后价，你的程序的“券后价”是用此值减优惠券金额得到的）
        # tk_total_commi => 647960       月支出佣金
        # tk_total_sales => 86240     淘客30天月推广量
        # ommission_rate 佣金
        # coupon_info 优惠券信息
        # item_url 产品地址
        # reserve_price 原价
        # shop_title 卖家昵称
        # title 产品标题
        # volume 30天销量
        # zk_final_price  商品折后价
        # tk_total_commi 月支出佣金
        # tk_total_sales  淘客30天月推广量
        # title = [
        #     "佣金", "优惠券信息", "产品地址", "原价", "卖家昵称", "产品标题", "30天销量", "商品折后价", "月支出佣金",
        #     "淘客30天月推广量"
        # ]
        # book = xlwt.Workbook()  # 创建一个excel对象
        # sheet = self.book.add_sheet('Sheet1', cell_overwrite_ok=True)  # 添加一个sheet页
        # sheet = self.excelFile.sheet_by_name('Sheet1')  # 获取一个sheet
        # for i in range(len(title)):  # 循环列
        #     sheet.write(0, i, title[i])  # 将title数组中的字段写入到0行i列中
        row = self.readExcelNrows()  # self.sheet.nrows
        # print('row %d' % row)
        for line in items:  # 循环字典
            # print(int(line['commission_rate']) / 100)
            # row = int(line) # 行
            commission_rate = int(line['commission_rate']) / 100
            self.sheet.write(row, 0, "%.02f%s" %
                             (commission_rate, '%'))  # commission_rate
            self.sheet.write(row, 1, line['coupon_info'])  # coupon_info
            self.sheet.write(row, 2, line['item_url'])  # item_url

            self.sheet.write(row, 3,
                             "%s元" % line['reserve_price'])  # reserve_price
            self.sheet.write(row, 4, line['shop_title'])  # shop_title
            self.sheet.write(row, 5, line['title'])  # title

            self.sheet.write(row, 6, line['volume'])  # volume
            self.sheet.write(row, 7, line['zk_final_price'])  # zk_final_price
            self.sheet.write(row, 8,
                             float(line['tk_total_commi']))  # tk_total_commi

            self.sheet.write(row, 9, line['tk_total_sales'])  # tk_total_sales
            row += 1
        self.saveExcel()
        # if row > 100:
        #     self.crateExcel2()  #文件太多了  分成多个文件
        # detester = time.strftime("%Y-%m-%d", time.localtime())
        # self.book.save(("%s_导出数据.xlsx" % detester))


# if __name__ == '__main__':
#     excelUtil =  ExcelUtil()
#     a = readExcel('testJson.json')
#     print(a)
#     writeM(a)