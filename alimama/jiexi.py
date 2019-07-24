# !/usr/bin/python3
import json
import time

# import xlwt
# import xlrd
from openpyxl import Workbook

title = [
    "佣金", "优惠券信息", "产品地址", "原价", "卖家昵称", "产品标题", "30天销量", "商品折后价", "月支出佣金",
    "淘客30天月推广量"
]
SHEET_NAME = 'Sheet1'


class ExcelUtil:
    def __init__(self, dirPath, fileName):
        self.dirPath = dirPath
        self.fileName = fileName
        self.fileCount = 1
        detester = time.strftime("%Y-%m-%d", time.localtime())
        self.path = "%s/%s%s_导出数据.xlsx" % (self.dirPath, self.fileName,
                                           detester)
        self.excelFile = self.createExcel(self.path)  # 创建一个excel

    def crateExcel2(self):
        self.saveExcel()
        #创建一个文件
        detester = time.strftime("%Y-%m-%d", time.localtime())
        self.path = "%s/%s%s_导出数据_%d.xlsx" % (self.dirPath, self.fileName,
                                              detester, self.fileCount)
        self.excelFile = self.createExcel(self.path)  # 创建一个excel
        self.fileCount += 1

    def readExcelNrows(self):
        try:
            # self.saveExcel()
            # book = xlrd.open_workbook(self.path)  # 打开一个excel
            # sheet = book.sheet_by_name(SHEET_NAME)  # 根据顺序获取sheet
            # return sheet.nrows
            #获取所有sheet页名字
            #sheet = self.excelFile.get_sheet_by_name(SHEET_NAME)
            row = self.sheet.max_row
            # print(row)
            return row
        except BaseException:
            return 1
            pass
        # return 1

    def readExcel(self, fileName):
        with open(fileName, 'r', encoding='utf8') as fr:
            data = json.load(fr)  # 用json中的load方法，将json串转换成字典
        return data

    def saveExcel(self):
        #detester = time.strftime("%Y-%m-%d", time.localtime())
        # self.excelFile.save(
        #     ("%s%s%s_导出数据.xlsx" % (self.dirPath, self.fileName, detester)))
        # print(self.dirPath)
        # self.excelFile.save(self.path)
        try:
            self.excelFile.save(self.path)
            # print(self.path)
        except BaseException:
            # print(" saveExcel BaseException")
            # print("Unexpected error:", sys.exc_info()[0])
            pass

    def createExcel(self, fileName):
        book = Workbook()  #创建文件对象
        # book = xlwt.Workbook()  # 创建一个excel对象 只写
        # self.sheet = book.add_sheet(SHEET_NAME,
        #                             cell_overwrite_ok=True)  # 添加一个sheet页
        self.sheet = book.active
        # self.sheet = book.create_sheet(SHEET_NAME)
        for i in range(len(title)):  # 循环列
            self.sheet.cell(1, i + 1, title[i])  # 将title数组中的字段写入到0行i列中
        # detester = time.strftime("%Y-%m-%d", time.localtime())

        book.save(fileName)

        return book
        # self.excelFile.save(
        #     ("%s/%s%s_导出数据.xlsx" % (self.dirPath, self.fileName, detester)))

    def writeM(self, items):
        # try:
        row = self.readExcelNrows()  # self.sheet.nrows
        # print('row %d' % row)
        for line in items:  # 循环字典
            # print(int(line['commission_rate']) / 100)
            # row = int(line) # 行
            commission_rate = int(line['commission_rate']) / 100
            # self.sheet.cell(row, 1, "%.02f%s" %
            #                 (commission_rate, '%'))  # commission_rate
            # self.sheet.cell(row, 2, line['coupon_info'])  # coupon_info
            # self.sheet.cell(row, 3, line['item_url'])  # item_url

            # self.sheet.cell(row, 4,
            #                 "%s元" % line['reserve_price'])  # reserve_price
            # self.sheet.cell(row, 5, line['shop_title'])  # shop_title
            # self.sheet.cell(row, 6, line['title'])  # title

            # self.sheet.cell(row, 7, line['volume'])  # volume
            # self.sheet.cell(row, 8, line['zk_final_price'])  # zk_final_price
            # self.sheet.cell(row, 9,
            #                 float(line['tk_total_commi']))  # tk_total_commi
            # self.sheet.cell(row, 10, line['tk_total_sales'])  # tk_total_sales
            # row += 1
            self.sheet.append([
                "%.02f%s" % (commission_rate, '%'), line['coupon_info'],
                line['item_url'],
                "%s元" % line['reserve_price'], line['shop_title'],
                line['title'], line['volume'], line['zk_final_price'],
                float(line['tk_total_commi']), line['tk_total_sales']
            ])
        # self.saveExcel()
        row = self.readExcelNrows()  # self.sheet.nrows
        if row >= 1000000:
            self.crateExcel2()  #文件太多了  分成多个文件
        # except BaseException:
        #     pass
        # detester = time.strftime("%Y-%m-%d", time.localtime())
        # self.book.save(("%s_导出数据.xlsx" % detester))


# if __name__ == '__main__':
# excelUtil =  ExcelUtil('/Users/twt/Documents/郭海山/API数据提取软件','test')
# a = excelUtil.readExcel('testJson.json')
# print(a)
# for row in range(70000):
#     excelUtil.writeM(a['result_list'])
# import openpyxl
# import time

# def saveItem(sheet, items):

#     for line in items:  # 循环字典
#         # print(int(line['commission_rate']) / 100)
#         # row = int(line) # 行

#         commission_rate = int(line['commission_rate']) / 100
#         # sheet.cell(row, 1,
#         #            "%.02f%s" % (commission_rate, '%'))  # commission_rate
#         # sheet.cell(row, 2, line['coupon_info'])  # coupon_info
#         # sheet.cell(row, 3, line['item_url'])  # item_url

#         # sheet.cell(row, 4, "%s元" % line['reserve_price'])  # reserve_price
#         # sheet.cell(row, 5, line['shop_title'])  # shop_title
#         # sheet.cell(row, 6, line['title'])  # title

#         # sheet.cell(row, 7, line['volume'])  # volume
#         # sheet.cell(row, 8, line['zk_final_price'])  # zk_final_price
#         # sheet.cell(row, 9, float(line['tk_total_commi']))  # tk_total_commi
#         # sheet.cell(row, 10, line['tk_total_sales'])  # tk_total_sales

#         sheet1.append([
#             "%.02f%s" % (commission_rate, '%'), line['coupon_info'],
#             line['item_url'],
#             "%s元" % line['reserve_price'], line['shop_title'], line['title'],
#             line['volume'], line['zk_final_price'],
#             float(line['tk_total_commi']), line['tk_total_sales']
#         ])
# f.save('openpyxlw.xlsx')

# time.perf_counter()
# f = openpyxl.Workbook()
# sheet1 = f.active
# with open('testJson.json', 'r', encoding='utf8') as fr:
#     data = json.load(fr)  # 用json中的load方法，将json串转换成字典
# for i in range(7553):
#     # sheet1.cell(row = i+1,column = 1,value = i)
#     # sheet1.append([i])
#     saveItem(sheet1, data['result_list'])
#     # sheet1.appand()
# f.save('openpyxlw.xlsx')
# t = time.perf_counter()
# print(t)
